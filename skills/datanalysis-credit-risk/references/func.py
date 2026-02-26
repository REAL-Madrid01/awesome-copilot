"""数据处理函数模块"""
import pandas as pd
import numpy as np
import toad
from typing import List, Dict, Tuple
import tqdm
from datetime import datetime

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    HAS_OPENPYXL = True
except:
    HAS_OPENPYXL = False


def get_dataset(data_pth: str, date_colName: str, y_colName: str, 
                org_colName: str, data_encode: str, key_colNames: List[str],
                drop_colNames: List[str] = None,
                miss_vals: List[int] = None) -> pd.DataFrame:
    """加载并格式化数据
    
    Args:
        data_pth: 数据路径
        date_colName: 日期列名
        y_colName: 标签列名
        org_colName: 机构列名
        data_encode: 数据编码
        key_colNames: 主键列（去重用）
        drop_colNames: 指定删除的列
        miss_vals: 异常值列表（替换为NaN），默认[-1, -999, -1111]
    """
    if drop_colNames is None:
        drop_colNames = []
    if miss_vals is None:
        miss_vals = [-1, -999, -1111]
    
    # 多格式读取
    for fmt, reader in [('parquet', pd.read_parquet), ('csv', pd.read_csv), 
                         ('xlsx', pd.read_excel), ('pkl', pd.read_pickle)]:
        try:
            data = reader(data_pth)
            break
        except:
            continue
    
    # 异常值替换为NaN
    data.replace({v: np.nan for v in miss_vals}, inplace=True)
    
    # 去重与过滤
    data = data[data[y_colName].isin([0, 1])]
    data = data.drop_duplicates(subset=key_colNames)
    
    # 删除无效列
    data.drop(columns=[c for c in drop_colNames if c in data.columns], errors='ignore')
    data.drop(columns=[c for c in data.columns if data[c].nunique() <= 1], errors='ignore')
    
    # 重命名
    data.rename(columns={date_colName: 'new_date', y_colName: 'new_target', 
                         org_colName: 'new_org'}, inplace=True)
    data['new_date'] = data['new_date'].astype(str).str.replace('-', '', regex=False).str[:8]
    data['new_date_ym'] = data['new_date'].str[:6]
    
    return data


def org_analysis(data: pd.DataFrame, oos_orgs: List[str] = None) -> pd.DataFrame:
    """机构样本统计分析
    
    Args:
        data: 数据
        oos_orgs: 贷外机构列表，用于标识贷外样本
    """
    stat = data.groupby(['new_org', 'new_date_ym']).agg(
        单月坏样本数=('new_target', 'sum'),
        单月总样本数=('new_target', 'count'),
        单月坏样率=('new_target', 'mean')
    ).reset_index()
    
    # 累计统计
    stat['总坏样本数'] = stat.groupby('new_org')['单月坏样本数'].transform('sum')
    stat['总样本数'] = stat.groupby('new_org')['单月总样本数'].transform('sum')
    stat['总坏样率'] = stat['总坏样本数'] / stat['总样本数']
    
    # 标识是否为贷外机构
    if oos_orgs and len(oos_orgs) > 0:
        stat['样本类型'] = stat['new_org'].apply(lambda x: '贷外' if x in oos_orgs else '建模')
    else:
        stat['样本类型'] = '建模'
    
    stat = stat.rename(columns={'new_org': '机构', 'new_date_ym': '年月'})
    
    # 按样本类型排序（建模在前，贷外在后）
    stat = stat.sort_values(['样本类型', '机构', '年月'], ascending=[True, True, True])
    stat = stat.reset_index(drop=True)
    
    return stat[['机构', '年月', '单月坏样本数', '单月总样本数', '单月坏样率', '总坏样本数', '总样本数', '总坏样率', '样本类型']]


def missing_check(data: pd.DataFrame, channel: Dict[str, List[str]] = None) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """计算缺失率 - 包含整体和机构级别的缺失率
    
    Returns:
        miss_detail: 缺失率明细（格式：变量，整体，机构1，机构2，...，机构n）
        miss_ch: 整体缺失率（每个变量的整体缺失率）
    """
    miss_vals = [-1, -999, -1111]
    miss_ch = []
    
    # 排除非变量列：record_id, target, org_info等
    exclude_cols = ['new_date', 'new_date_ym', 'new_target', 'new_org', 'record_id', 'target', 'org_info']
    cols = [c for c in data.columns if c not in exclude_cols]
    
    # 计算整体缺失率
    for col in tqdm.tqdm(cols, desc="缺失率"):
        rate = ((data[col].isin(miss_vals)) | (data[col].isna())).mean()
        miss_ch.append({'变量': col, '整体缺失率': round(rate, 4)})
    
    miss_ch = pd.DataFrame(miss_ch)
    
    # 计算机构级别缺失率并转换为宽表格式
    orgs = sorted(data['new_org'].unique())
    miss_detail_dict = {'变量': []}
    miss_detail_dict['整体'] = []
    
    for org in orgs:
        miss_detail_dict[org] = []
    
    for col in cols:
        miss_detail_dict['变量'].append(col)
        # 整体缺失率
        overall_rate = ((data[col].isin(miss_vals)) | (data[col].isna())).mean()
        miss_detail_dict['整体'].append(round(overall_rate, 4))
        
        # 各机构缺失率
        for org in orgs:
            org_data = data[data['new_org'] == org]
            rate = ((org_data[col].isin(miss_vals)) | (org_data[col].isna())).mean()
            miss_detail_dict[org].append(round(rate, 4))
    
    miss_detail = pd.DataFrame(miss_detail_dict)
    # 按整体缺失率降序排序
    miss_detail = miss_detail.sort_values('整体', ascending=False)
    miss_detail = miss_detail.reset_index(drop=True)
    
    return miss_detail, miss_ch


def calculate_iv(data: pd.DataFrame, features: List[str], n_jobs: int = 4) -> pd.DataFrame:
    """计算IV值 - 使用toad.transform.Combiner分箱，箱子数设为5，保留NaN值"""
    import tqdm
    from joblib import Parallel, delayed
    
    def _calc_iv(f):
        try:
            # 使用 toad.transform.Combiner 分箱，箱子数设为5
            c = toad.transform.Combiner()
            data_temp = data[[f, 'new_target']].copy()
            data_temp.columns = ['x', 'y']
            data_temp['x_bin'] = c.fit_transform(X=data_temp['x'], y=data_temp['y'], method='dt', n_bins=5, min_samples=0.05/5, empty_separate=True)
            
            # 使用分箱后的数据计算IV值
            iv_df = toad.quality(data_temp[['x_bin', 'y']], 'y', iv_only=True)
            if 'iv' in iv_df.columns and len(iv_df) > 0:
                iv_value = iv_df['iv'].iloc[0]
                if not np.isnan(iv_value):
                    return {'变量': f, 'IV': round(iv_value, 4)}
            return None
        except Exception as e:
            print(f"   IV计算异常: 变量={f}, 错误={e}")
            return None
    
    # 使用tqdm显示进度
    results = Parallel(n_jobs=n_jobs, verbose=0)(
        delayed(_calc_iv)(f) for f in features
    )
    iv_list = [r for r in results if r is not None]
    
    if len(iv_list) == 0:
        print(f"   IV计算结果为空，特征数={len(features)}")
        return pd.DataFrame(columns=['变量', 'IV'])
    
    return pd.DataFrame(iv_list).sort_values('IV', ascending=False)


def calculate_corr(data: pd.DataFrame, features: List[str]) -> pd.DataFrame:
    """计算相关性矩阵"""
    corr = data[features].corr().abs()
    return corr


def export_report_xlsx(filepath: str, data_name: str, data: pd.DataFrame, 
                       sheet_name: str, description: str = ""):
    """导出xlsx报告 - 支持追加"""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(filepath)
        ws = wb.create_sheet(sheet_name)
    except:
        wb = Workbook()
        ws = wb.active
        ws.title = sheet_name
    
    # 写入描述
    ws['A1'] = f"数据: {data_name}"
    ws['A2'] = f"时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    if description:
        ws['A3'] = f"说明: {description}"
    
    # 写入数据
    start_row = 5
    for i, col in enumerate(data.columns):
        ws.cell(start_row, i+1, col)
    
    for i, row in enumerate(data.values):
        for j, val in enumerate(row):
            ws.cell(start_row+1+i, j+1, val)
    
    # 样式
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[start_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    wb.save(filepath)
    print(f"[{sheet_name}] 已保存到 {filepath}")
