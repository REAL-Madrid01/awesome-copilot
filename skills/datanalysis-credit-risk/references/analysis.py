"""变量筛选与分析模块 - 精简版
PSI计算在func.py中复用，analysis.py只做变量筛选
"""
import pandas as pd
import numpy as np
import toad
from typing import List, Dict, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime
import lightgbm as lgb
from sklearn.model_selection import train_test_split
from sklearn.metrics import roc_auc_score
from joblib import Parallel, delayed


def drop_abnormal_ym(data: pd.DataFrame, min_ym_bad_sample: int = 1, 
                     min_ym_sample: int = 500) -> tuple:
    """过滤异常月份 - 整体统计，不分机构"""
    stat = data.groupby('new_date_ym').agg(
        bad_cnt=('new_target', 'sum'),
        total=('new_target', 'count')
    ).reset_index()
    
    abnormal = stat[(stat['bad_cnt'] < min_ym_bad_sample) | (stat['total'] < min_ym_sample)]
    abnormal = abnormal.rename(columns={'new_date_ym': '年月'})
    abnormal['去除条件'] = abnormal.apply(
        lambda x: f'坏样本数{x["bad_cnt"]}小于{min_ym_bad_sample}' if x['bad_cnt'] < min_ym_bad_sample else f'总样本数{x["total"]}小于{min_ym_sample}', axis=1
    )
    
    if len(abnormal) > 0:
        data = data[~data['new_date_ym'].isin(abnormal['年月'])]
    
    # 移除空行
    abnormal = abnormal.dropna(how='all')
    abnormal = abnormal.reset_index(drop=True)
    
    return data, abnormal


def drop_highmiss_features(data: pd.DataFrame, miss_channel: pd.DataFrame, 
                           threshold: float = 0.6) -> tuple:
    """剔除高缺失变量"""
    high_miss = miss_channel[miss_channel['整体缺失率'] > threshold].copy()
    high_miss['缺失率'] = high_miss['整体缺失率']
    
    # 修改去除条件写法，显示具体缺失率值
    high_miss['去除条件'] = high_miss.apply(
        lambda x: f'该变量整体缺失率为{x["缺失率"]:.4f}，超过阈值{threshold}', axis=1
    )
    
    # 移除空行
    high_miss = high_miss.dropna(how='all')
    high_miss = high_miss.reset_index(drop=True)
    
    # 剔除高缺失变量
    if len(high_miss) > 0 and '变量' in high_miss.columns:
        to_drop = high_miss['变量'].tolist()
        data = data.drop(columns=[c for c in to_drop if c in data.columns])
    
    return data, high_miss[['变量', '缺失率', '去除条件']]


def drop_lowiv_features(data: pd.DataFrame, features: List[str], 
                       overall_iv_threshold: float = 0.05, org_iv_threshold: float = 0.02,
                       max_org_threshold: int = 8, n_jobs: int = 4) -> tuple:
    """剔除低IV变量 - 多进程版本，返回IV明细和IV处理表
    
    Args:
        overall_iv_threshold: 整体IV阈值，低于此值记录到IV处理表
        org_iv_threshold: 单机构IV阈值，低于此值认为不满足
        max_org_threshold: 最大容忍机构数，超过此数机构IV低于阈值则记录到IV处理表
    
    Returns:
        data: 剔除后的数据
        iv_detail: IV明细（每个变量在每个机构上以及整体上的IV值）
        iv_process: IV处理表（不满足设定条件的变量）
    """
    from references.func import calculate_iv
    from joblib import Parallel, delayed
    
    orgs = sorted(data['new_org'].unique())
    
    print(f"   IV计算: 特征数={len(features)}, 机构数={len(orgs)}")
    
    # 一次性计算所有机构的IV值
    def _calc_org_iv(org):
        org_data = data[data['new_org'] == org]
        org_iv = calculate_iv(org_data, features, n_jobs=1)
        if len(org_iv) > 0:
            org_iv = org_iv.rename(columns={'IV': 'IV值'})
            org_iv['机构'] = org
            return org_iv
        return None
    
    # 计算整体IV
    print(f"   计算整体IV...")
    iv_overall = calculate_iv(data, features, n_jobs=n_jobs)
    print(f"   整体IV计算结果: {len(iv_overall)}个变量")
    if len(iv_overall) == 0:
        print(f"   警告: 整体IV计算结果为空，返回空表")
        return data, pd.DataFrame(columns=['变量', 'IV值', '机构', '类型']), pd.DataFrame(columns=['变量', '整体IV', '低IV机构数', '处理原因'])
    iv_overall = iv_overall.rename(columns={'IV': 'IV值'})
    
    # 并行计算所有机构的IV值
    print(f"   并行计算{len(orgs)}个机构的IV值...")
    iv_by_org_results = Parallel(n_jobs=n_jobs, verbose=0)(
        delayed(_calc_org_iv)(org) for org in orgs
    )
    iv_by_org = [r for r in iv_by_org_results if r is not None]
    iv_by_org = pd.concat(iv_by_org, ignore_index=True) if iv_by_org else pd.DataFrame(columns=['变量', 'IV值', '机构'])
    print(f"   分机构IV汇总: {len(iv_by_org)}条记录")
    
    # 转换为宽表格式：变量，整体，机构1，机构2，...，机构n
    iv_detail_dict = {'变量': []}
    iv_detail_dict['整体'] = []
    
    for org in orgs:
        iv_detail_dict[org] = []
    
    # 获取所有变量
    all_vars = set(iv_overall['变量'].tolist())
    if len(iv_by_org) > 0:
        all_vars.update(iv_by_org['变量'].tolist())
    all_vars = sorted(all_vars)
    
    for var in all_vars:
        iv_detail_dict['变量'].append(var)
        
        # 整体IV
        var_overall = iv_overall[iv_overall['变量'] == var]
        if len(var_overall) > 0:
            iv_detail_dict['整体'].append(var_overall['IV值'].values[0])
        else:
            iv_detail_dict['整体'].append(None)
        
        # 各机构IV
        for org in orgs:
            var_org = iv_by_org[iv_by_org['机构'] == org]
            var_org = var_org[var_org['变量'] == var]
            if len(var_org) > 0:
                iv_detail_dict[org].append(var_org['IV值'].values[0])
            else:
                iv_detail_dict[org].append(None)
    
    iv_detail = pd.DataFrame(iv_detail_dict)
    # 按照整体IV降序排序
    iv_detail = iv_detail.sort_values('整体', ascending=False)
    iv_detail = iv_detail.reset_index(drop=True)
    
    # 标记不满足条件的变量
    # 1. 整体IV低于阈值
    iv_overall_low = iv_overall[iv_overall['IV值'] < overall_iv_threshold]['变量'].tolist()
    
    # 2. 单机构IV低于阈值的机构数
    if len(iv_by_org) > 0:
        iv_by_org_low = iv_by_org[iv_by_org['IV值'] < org_iv_threshold].groupby('变量').size().reset_index()
        iv_by_org_low.columns = ['变量', '低IV机构数']
    else:
        iv_by_org_low = pd.DataFrame(columns=['变量', '低IV机构数'])
    
    # 获取每个变量的低IV机构列表
    low_iv_orgs_dict = {}
    if len(iv_by_org) > 0:
        for var in iv_by_org['变量'].unique():
            var_orgs = iv_by_org[(iv_by_org['变量'] == var) & (iv_by_org['IV值'] < org_iv_threshold)]['机构'].tolist()
            low_iv_orgs_dict[var] = var_orgs
    
    # 3. 标记需要处理的变量
    iv_process = []
    
    # 调试信息：统计IV分布
    if len(iv_overall) > 0:
        print(f"   整体IV统计: 最小值={iv_overall['IV值'].min():.4f}, 最大值={iv_overall['IV值'].max():.4f}, 中位数={iv_overall['IV值'].median():.4f}")
        print(f"   整体IV小于{overall_iv_threshold}的变量数: {(iv_overall['IV值'] < overall_iv_threshold).sum()}/{len(iv_overall)}")
    
    if len(iv_by_org_low) > 0:
        print(f"   分机构IV小于{org_iv_threshold}的变量统计:")
        print(f"     最多低IV机构数: {iv_by_org_low['低IV机构数'].max()}")
        print(f"     低IV机构数大于等于{max_org_threshold}的变量数: {(iv_by_org_low['低IV机构数'] >= max_org_threshold).sum()}/{len(iv_by_org_low)}")
    
    for var in features:
        reasons = []
        
        # 检查整体IV
        var_overall_iv = iv_overall[iv_overall['变量'] == var]['IV值'].values
        if len(var_overall_iv) > 0 and var_overall_iv[0] < overall_iv_threshold:
            reasons.append(f'整体IV{var_overall_iv[0]:.4f}小于阈值{overall_iv_threshold}')
        
        # 检查分机构IV
        var_org_low = iv_by_org_low[iv_by_org_low['变量'] == var]
        if len(var_org_low) > 0 and var_org_low['低IV机构数'].values[0] >= max_org_threshold:
            reasons.append(f'在{var_org_low["低IV机构数"].values[0]}个机构上IV小于阈值{org_iv_threshold}')
        
        if reasons:
            iv_process.append({
                '变量': var,
                '处理原因': '; '.join(reasons),
                '低IV机构': ','.join(low_iv_orgs_dict.get(var, []))
            })
    
    iv_process = pd.DataFrame(iv_process)
    iv_process = iv_process.reset_index(drop=True)
    
    # 剔除不满足条件的变量
    if len(iv_process) > 0 and '变量' in iv_process.columns:
        to_drop = iv_process['变量'].tolist()
        data = data.drop(columns=[c for c in to_drop if c in data.columns])
    
    return data, iv_detail, iv_process


def drop_highcorr_features(data: pd.DataFrame, features: List[str],
                           threshold: float = 0.8, gain_dict: dict = None, top_n_keep: int = 20) -> tuple:
    """剔除高相关变量 - 基于原始gain，每次剔除一个变量
    
    Args:
        data: 数据
        features: 特征列表
        threshold: 相关性阈值
        gain_dict: 变量到原始gain的映射字典
        top_n_keep: 保留原始gain排名前N的变量
    
    Returns:
        data: 剔除后的数据
        dropped_info: 剔除信息
    """
    if gain_dict is None:
        gain_dict = {}
    
    # 获取当前特征列表（只存在于数据中的特征）
    current_features = [f for f in features if f in data.columns]
    
    if len(current_features) == 0:
        return data, pd.DataFrame(columns=['变量', '相关变量', '去除条件'])
    
    # 确定需要保留的变量（原始gain排名前N）
    if gain_dict:
        # 只考虑存在于当前特征中的变量
        current_gain_dict = {k: v for k, v in gain_dict.items() if k in current_features}
        if current_gain_dict:
            sorted_features = sorted(current_gain_dict.keys(), key=lambda x: current_gain_dict[x], reverse=True)
            top_features = set(sorted_features[:top_n_keep])
            # 创建变量到排名的映射
            rank_dict = {v: i+1 for i, v in enumerate(sorted_features)}
        else:
            top_features = set()
            rank_dict = {}
    else:
        top_features = set()
        rank_dict = {}
    
    dropped_info = []
    
    # 循环剔除，直到没有高相关变量对
    while True:
        # 重新计算相关性矩阵（只针对当前剩余的特征）
        current_features = [f for f in current_features if f in data.columns]
        if len(current_features) < 2:
            break
        
        corr = data[current_features].corr().abs()
        upper = corr.where(np.triu(np.ones(corr.shape), k=1).astype(bool))
        
        # 找到所有高相关变量对
        high_corr_pairs = []
        for i, col1 in enumerate(upper.columns):
            for col2 in upper.columns[i+1:]:
                corr_val = upper.loc[col1, col2]
                if pd.notna(corr_val) and corr_val > threshold:
                    high_corr_pairs.append((col1, col2, corr_val))
        
        if not high_corr_pairs:
            break
        
        # 对于每个高相关变量对，选择原始gain较小的变量作为候选剔除
        candidates = set()
        for col1, col2, corr_val in high_corr_pairs:
            # 跳过top N保留的变量
            if col1 in top_features and col2 in top_features:
                continue
            
            gain1 = gain_dict.get(col1, 0)
            gain2 = gain_dict.get(col2, 0)
            
            # 选择原始gain较小的变量
            if gain1 <= gain2:
                candidates.add(col1)
            else:
                candidates.add(col2)
        
        if not candidates:
            break
        
        # 在候选变量中选择原始gain最小的进行剔除
        candidates_list = list(candidates)
        candidates_with_gain = [(c, gain_dict.get(c, 0)) for c in candidates_list]
        candidates_with_gain.sort(key=lambda x: x[1])
        to_drop = candidates_with_gain[0][0]
        
        # 找到与该变量高相关的所有变量
        related_vars = []
        for col1, col2, corr_val in high_corr_pairs:
            if col1 == to_drop:
                related_vars.append((col2, corr_val))
            elif col2 == to_drop:
                related_vars.append((col1, corr_val))
        
        # 记录剔除信息
        # 相关变量列：显示变量名和相似度值（相关性值）
        related_str = ','.join([f"{v}(相似度={c:.4f})" for v, c in related_vars])
        # 去除条件列：显示相关变量及其对应的gain值
        gain_str = ','.join([f"{v}(gain={gain_dict.get(v, 0):.2f})" for v, c in related_vars])
        dropped_info.append({
            '变量': to_drop,
            '原始gain': gain_dict.get(to_drop, 0),
            '原始gain排名': rank_dict.get(to_drop, '-'),
            '相关变量': related_str,
            '去除条件': gain_str
        })
        
        # 从数据中删除该变量
        data = data.drop(columns=[to_drop], errors='ignore')
        current_features.remove(to_drop)
        
        print(f"   剔除变量: {to_drop} (原始gain={gain_dict.get(to_drop, 0):.2f})")
    
    # 转换为DataFrame并按原始gain降序排序
    dropped_df = pd.DataFrame(dropped_info)
    if len(dropped_df) > 0:
        dropped_df = dropped_df.sort_values('原始gain', ascending=False)
        dropped_df = dropped_df.reset_index(drop=True)
    
    return data, dropped_df


def drop_highnoise_features(data: pd.DataFrame, features: List[str],
                           n_estimators: int = 100, max_depth: int = 5, gain_threshold: float = 50) -> tuple:
    """Null Importance去除高噪音变量"""
    # 检查特征列表是否为空
    if len(features) == 0:
        print("   没有特征需要处理")
        return data, pd.DataFrame(columns=['变量', '原始gain', '反转后gain'])
    
    # 检查数据是否足够
    if len(data) < 1000:
        print(f"   数据量不足({len(data)}条)，跳过Null Importance")
        return data, pd.DataFrame(columns=['变量', '原始gain', '反转后gain'])
    
    X = data[features].copy()
    Y = data['new_target'].copy()
    
    # 检查X是否为空或包含NaN
    if X.shape[1] == 0:
        print("   特征数据为空，跳过Null Importance")
        return data, pd.DataFrame(columns=['变量', '原始gain', '反转后gain'])
    
    # 填充NaN
    X = X.fillna(0)
    
    # 打乱标签
    Y_permuted = Y.copy()
    for _ in range(20):
        Y_permuted = np.random.permutation(Y_permuted)
    
    clf = lgb.LGBMClassifier(
        objective='binary', boosting_type='gbdt', learning_rate=0.05,
        max_depth=max_depth, min_child_samples=2000, min_child_weight=20,
        n_estimators=n_estimators, num_leaves=2**max_depth - 1, n_jobs=-1, verbose=-1
    )
    
    clf_permuted = lgb.LGBMClassifier(
        objective='binary', boosting_type='gbdt', learning_rate=0.05,
        max_depth=max_depth, min_child_samples=2000, min_child_weight=20,
        n_estimators=n_estimators, num_leaves=2**max_depth - 1, n_jobs=-1, verbose=-1
    )
    
    results, results_permuted = [], []
    
    print("Null Importance计算中...")
    for i in range(2):
        random_n = np.random.randint(30)
        
        X_train, X_test, y_train, y_test = train_test_split(X, Y, test_size=0.3, random_state=random_n)
        
        # 检查训练数据是否有效
        if X_train.shape[0] == 0 or X_test.shape[0] == 0:
            print(f"  轮次{i+1}: 数据分割失败，跳过")
            continue
        
        clf.fit(X_train, y_train)
        
        X_train_, X_test_, y_train_, y_test_ = train_test_split(X, Y_permuted, test_size=0.3, random_state=random_n)
        
        if X_train_.shape[0] == 0 or X_test_.shape[0] == 0:
            print(f"  轮次{i+1}: 打乱数据分割失败，跳过")
            continue
        
        clf_permuted.fit(X_train_, y_train_)
        
        imp_real = pd.DataFrame({
            'feature': clf.booster_.feature_name(),
            'gain': clf.booster_.feature_importance(importance_type='gain')
        })
        imp_perm = pd.DataFrame({
            'feature': clf_permuted.booster_.feature_name(),
            'gain': clf_permuted.booster_.feature_importance(importance_type='gain')
        })
        
        results.append(imp_real)
        results_permuted.append(imp_perm)
        
        train_auc = roc_auc_score(y_train, clf.predict_proba(X_train)[:, 1])
        test_auc = roc_auc_score(y_test, clf.predict_proba(X_test)[:, 1])
        print(f"  轮次{i+1}: train_auc={train_auc:.3f}, test_auc={test_auc:.3f}")
    
    # 检查是否有有效结果
    if len(results) == 0 or len(results_permuted) == 0:
        print("   没有有效的训练结果，跳过Null Importance")
        return data, pd.DataFrame(columns=['变量', '原始gain', '反转后gain'])
    
    imp_real_avg = pd.concat(results).groupby('feature')['gain'].mean().reset_index()
    imp_perm_avg = pd.concat(results_permuted).groupby('feature')['gain'].mean().reset_index()
    
    comparison = imp_real_avg.merge(imp_perm_avg, on='feature', suffixes=('_real', '_perm'))
    comparison['gain_real'] = comparison['gain_real'].fillna(0)
    comparison['gain_perm'] = comparison['gain_perm'].fillna(0)
    
    # 使用反转前后gain值差别绝对值小于50的条件
    comparison['gain_diff'] = (comparison['gain_real'] - comparison['gain_perm']).abs()
    noise_features = comparison[comparison['gain_diff'] < gain_threshold]['feature'].tolist()
    
    # 列出所有变量的原始gain和反转后gain
    dropped_info = pd.DataFrame({
        '变量': comparison['feature'].values,
        '原始gain': comparison['gain_real'].values,
        '反转后gain': comparison['gain_perm'].values
    })
    # 增加状态列，去除的要标明去除，不去除的标明保留
    dropped_info['状态'] = dropped_info.apply(
        lambda x: '去除' if np.abs(x['原始gain'] - x['反转后gain']) < gain_threshold else '保留', axis=1
    )
    # 按原始gain降序排序
    dropped_info = dropped_info.sort_values('原始gain', ascending=False)
    dropped_info = dropped_info.reset_index(drop=True)
    # 添加原始gain排名列
    dropped_info['原始gain排名'] = range(1, len(dropped_info) + 1)
    
    data = data.drop(columns=[c for c in noise_features if c in data.columns])
    
    print(f"  剔除{len(noise_features)}个噪音变量")
    return data, dropped_info


def _calc_single_psi(args):
    """计算单个机构单个变量的PSI - NaN作为单独箱子"""
    org, train_month, test_month, train_n, test_n, f, data_ref, min_sample = args
    
    try:
        org_data = data_ref[data_ref['new_org'] == org]
        train_data = org_data[org_data['new_date_ym'] == train_month]
        test_data = org_data[org_data['new_date_ym'] == test_month]
        
        # 获取数据
        train_vals = train_data[f].values
        test_vals = test_data[f].values
        
        # 标记NaN
        train_nan_mask = pd.isna(train_vals)
        test_nan_mask = pd.isna(test_vals)
        
        # 非NaN值用于分箱
        train_nonan = train_vals[~train_nan_mask]
        test_nonan = test_vals[~test_nan_mask]
        
        if len(train_nonan) < min_sample or len(test_nonan) < min_sample:
            return {
                '机构': org, '日期': f"{train_month}->{test_month}",
                '变量': f, 'PSI': None, '有效计算': 0,
                '样本数': train_n
            }
        
        # 基于非NaN数据分箱（10分箱）
        try:
            bins = pd.qcut(train_nonan, q=10, duplicates='drop', retbins=True)[1]
        except:
            bins = pd.cut(train_nonan, bins=10, retbins=True)[1]
        
        # 计算各箱子占比（包含NaN箱子）
        train_counts = []
        test_counts = []
        
        for i in range(len(bins)):
            if i == 0:
                train_counts.append((~train_nan_mask & (train_vals <= bins[i])).sum())
                test_counts.append((~test_nan_mask & (test_vals <= bins[i])).sum())
            else:
                train_counts.append((~train_nan_mask & (train_vals > bins[i-1]) & (train_vals <= bins[i])).sum())
                test_counts.append((~test_nan_mask & (test_vals > bins[i-1]) & (test_vals <= bins[i])).sum())
        
        # NaN箱子
        train_counts.append(train_nan_mask.sum())
        test_counts.append(test_nan_mask.sum())
        
        # 转为占比
        train_pct = np.array(train_counts) / len(train_vals)
        test_pct = np.array(test_counts) / len(test_vals)
        
        # 避免0值
        train_pct = np.where(train_pct == 0, 1e-6, train_pct)
        test_pct = np.where(test_pct == 0, 1e-6, test_pct)
        
        # 计算PSI
        psi = np.sum((test_pct - train_pct) * np.log(test_pct / train_pct))
        
        return {
            '机构': org, '日期': f"{train_month}->{test_month}",
            '变量': f, 'PSI': round(psi, 4), '有效计算': 1,
            '样本数': train_n
        }
    except Exception as e:
        return {
            '机构': org, '日期': f"{train_month}->{test_month}",
            '变量': f, 'PSI': None, '有效计算': 0,
            '样本数': train_n
        }


def drop_highpsi_features(data: pd.DataFrame, features: List[str],
                         psi_threshold: float = 0.1, max_months_ratio: float = 1/3,
                         max_orgs: int = 4, min_sample_per_month: int = 100, n_jobs: int = 4) -> tuple:
    """剔除高PSI变量 - 分机构+逐月份版本
    
    多进程在变量层面，每个机构循环，机构内变量并行计算
    
    Args:
        psi_threshold: PSI阈值，超过此值认为不稳定
        max_months_ratio: 最大容忍月份比例，超过此比例月份PSI超过阈值则记录到处理表
        max_orgs: 最大容忍机构数，超过此数机构不稳定则记录到处理表
        min_sample_per_month: 每月最小样本数
    
    Returns:
        data: 剔除后的数据
        psi_detail: PSI明细（每个变量在每个机构每个月上的PSI值）
        psi_process: PSI处理表（不满足设定条件的变量）
    """
    orgs = data['new_org'].unique()
    
    # 构建任务列表：每个机构、每对月份、每个变量
    tasks = []
    for org in orgs:
        org_data = data[data['new_org'] == org]
        months = sorted(org_data['new_date_ym'].unique())
        
        if len(months) < 2:
            continue
        
        for i in range(len(months) - 1):
            train_month = months[i]
            test_month = months[i + 1]
            
            train_data = org_data[org_data['new_date_ym'] == train_month]
            test_data = org_data[org_data['new_date_ym'] == test_month]
            
            train_n = len(train_data)
            test_n = len(test_data)
            
            for f in features:
                tasks.append((org, train_month, test_month, train_n, test_n, f, data, min_sample_per_month))
    
    # 多进程计算PSI（变量层面并行）
    print(f"   PSI计算: {len(tasks)}个任务, 使用{n_jobs}进程")
    results = Parallel(n_jobs=n_jobs, verbose=0)(delayed(_calc_single_psi)(task) for task in tasks)
    
    psi_df = pd.DataFrame(results)
    
    if len(psi_df) == 0:
        return data, pd.DataFrame(columns=['变量', '机构', '月份', 'PSI值']), pd.DataFrame(columns=['变量', '处理原因'])
    
    # 筛选有效计算的记录
    valid_psi = psi_df[psi_df['有效计算'] == 1].copy()
    
    if len(valid_psi) == 0:
        return data, pd.DataFrame(columns=['变量', '机构', '月份', 'PSI值']), pd.DataFrame(columns=['变量', '处理原因'])
    
    # PSI明细表：每个变量在每个机构每个月上的PSI值
    # 日期改为单个月份，初始月份PSI值为0
    psi_detail = valid_psi[['机构', '日期', '变量', 'PSI']].copy()
    
    # 解析日期，提取测试月份
    psi_detail['月份'] = psi_detail['日期'].apply(lambda x: x.split('->')[1] if '->' in x else x)
    psi_detail = psi_detail.rename(columns={'PSI': 'PSI值'})
    
    # 按照变量，机构，月份升序排序
    psi_detail = psi_detail.sort_values(['变量', '机构', '月份'], ascending=[True, True, True])
    
    # 获取所有机构和月份
    all_orgs = sorted(psi_detail['机构'].unique())
    all_vars = sorted(psi_detail['变量'].unique())
    
    # 构建完整的PSI明细表（包含初始月份，PSI值为0）
    psi_detail_list = []
    for org in all_orgs:
        org_data = psi_detail[psi_detail['机构'] == org]
        if len(org_data) == 0:
            continue
        
        # 获取该机构的所有月份
        months = sorted(org_data['月份'].unique())
        
        for var in all_vars:
            var_data = org_data[org_data['变量'] == var]
            if len(var_data) == 0:
                continue
            
            # 初始月份PSI值为0
            psi_detail_list.append({
                '机构': org,
                '变量': var,
                '月份': months[0],
                'PSI值': 0.0
            })
            
            # 后续月份PSI值为计算结果
            for i in range(1, len(months)):
                month = months[i]
                var_month_data = var_data[var_data['月份'] == month]
                if len(var_month_data) > 0:
                    psi_value = var_month_data['PSI值'].values[0]
                else:
                    psi_value = 0.0
                psi_detail_list.append({
                    '机构': org,
                    '变量': var,
                    '月份': month,
                    'PSI值': psi_value
                })
    
    psi_detail = pd.DataFrame(psi_detail_list)
    psi_detail = psi_detail[['机构', '变量', '月份', 'PSI值']]
    psi_detail = psi_detail.reset_index(drop=True)
    # 按照变量，机构，月份升序排序
    psi_detail = psi_detail.sort_values(['变量', '机构', '月份'], ascending=[True, True, True])
    psi_detail = psi_detail.reset_index(drop=True)
    
    # 标记不稳定
    valid_psi['不稳定'] = (valid_psi['PSI'] > psi_threshold).astype(int)
    
    # 汇总：每个机构每个变量的不稳定月份数和总月份数
    org_summary = valid_psi.groupby(['机构', '变量']).agg(
        不稳定月份数=('不稳定', 'sum'),
        总月份数=('变量', 'count')
    ).reset_index()
    
    # 标记每个机构每个变量是否不稳定
    # 确保阈值至少为1，避免机构月份数少时过于严格
    org_summary['不稳定阈值'] = org_summary['总月份数'].apply(
        lambda x: max(1, int(x * max_months_ratio))
    )
    org_summary['是否不稳定'] = org_summary['不稳定月份数'] >= org_summary['不稳定阈值']
    
    # 机构层面汇总：不稳定的机构数
    org_count = len(orgs)
    channel_summary = org_summary.groupby('变量').apply(
        lambda x: pd.Series({
            '机构数': org_count,
            '不稳定机构数': x['是否不稳定'].sum()
        })
    ).reset_index()
    
    # 标记需要处理的变量
    channel_summary['需处理'] = channel_summary['不稳定机构数'] >= max_orgs
    channel_summary['处理原因'] = channel_summary.apply(
        lambda x: f'在{x["不稳定机构数"]}个机构上PSI不稳定' if x['需处理'] else '', axis=1
    )
    
    # 获取每个变量的不稳定机构列表
    unstable_orgs_dict = {}
    for var in org_summary['变量'].unique():
        var_orgs = org_summary[(org_summary['变量'] == var) & (org_summary['是否不稳定'] == True)]['机构'].tolist()
        unstable_orgs_dict[var] = var_orgs
    
    # PSI处理表：不满足设定条件的变量
    psi_process = channel_summary[channel_summary['需处理']].copy()
    psi_process['不稳定机构'] = psi_process['变量'].apply(lambda x: ','.join(unstable_orgs_dict.get(x, [])))
    psi_process = psi_process[['变量', '处理原因', '不稳定机构']]
    psi_process = psi_process.reset_index(drop=True)
    
    # 筛选要剔除的变量
    if len(psi_process) > 0 and '变量' in psi_process.columns:
        to_drop_vars = psi_process['变量'].tolist()
        data = data.drop(columns=[c for c in to_drop_vars if c in data.columns])
    
    return data, psi_detail, psi_process


def iv_distribution_by_org(iv_detail: pd.DataFrame, oos_orgs: list = None, iv_bins: list = [0, 0.02, 0.05, 0.1, float('inf')]) -> pd.DataFrame:
    """统计每个机构下不同IV区间的变量个数和占比
    
    Args:
        iv_detail: IV明细表（包含变量、整体、各机构列）
        oos_orgs: 贷外机构列表
        iv_bins: IV区间边界 [0, 0.02, 0.05, 0.1, inf]
    
    Returns:
        IV分布统计表
    """
    if oos_orgs is None:
        oos_orgs = []
    
    # 获取机构列（排除'变量'和'整体'列）
    org_cols = [c for c in iv_detail.columns if c not in ['变量', '整体']]
    
    # 定义区间标签
    bin_labels = ['[0, 0.02)', '[0.02, 0.05)', '[0.05, 0.1)', '[0.1, +∞)']
    
    result = []
    
    # 各机构统计（不计算整体）
    for org in org_cols:
        org_iv = iv_detail[org].dropna()
        total_vars = len(org_iv)
        
        # 判断机构类型
        org_type = '贷外' if org in oos_orgs else '建模'
        
        for i in range(len(iv_bins) - 1):
            lower = iv_bins[i]
            upper = iv_bins[i + 1]
            if upper == float('inf'):
                count = ((org_iv >= lower)).sum()
            else:
                count = ((org_iv >= lower) & (org_iv < upper)).sum()
            ratio = count / total_vars if total_vars > 0 else 0
            result.append({
                '机构': org,
                '类型': org_type,
                'IV区间': bin_labels[i],
                '变量个数': count,
                '占比': f'{ratio:.2%}'
            })
    
    return pd.DataFrame(result)


def psi_distribution_by_org(psi_detail: pd.DataFrame, oos_orgs: list = None, psi_bins: list = [0, 0.05, 0.1, float('inf')]) -> pd.DataFrame:
    """统计每个机构下不同PSI区间的变量个数和占比
    
    Args:
        psi_detail: PSI明细表（包含机构、变量、月份、PSI值列）
        oos_orgs: 贷外机构列表
        psi_bins: PSI区间边界 [0, 0.05, 0.1, inf]
    
    Returns:
        PSI分布统计表
    """
    if oos_orgs is None:
        oos_orgs = []
    
    # 定义区间标签
    bin_labels = ['[0, 0.05)', '[0.05, 0.1)', '[0.1, +∞)']
    
    result = []
    
    # 获取所有机构
    orgs = psi_detail['机构'].unique()
    
    for org in orgs:
        org_data = psi_detail[psi_detail['机构'] == org]
        
        # 判断机构类型
        org_type = '贷外' if org in oos_orgs else '建模'
        
        # 对每个变量，取其最大PSI值
        var_max_psi = org_data.groupby('变量')['PSI值'].max()
        total_vars = len(var_max_psi)
        
        for i in range(len(psi_bins) - 1):
            lower = psi_bins[i]
            upper = psi_bins[i + 1]
            if upper == float('inf'):
                count = ((var_max_psi >= lower)).sum()
            else:
                count = ((var_max_psi >= lower) & (var_max_psi < upper)).sum()
            ratio = count / total_vars if total_vars > 0 else 0
            result.append({
                '机构': org,
                '类型': org_type,
                'PSI区间': bin_labels[i],
                '变量个数': count,
                '占比': f'{ratio:.2%}'
            })
    
    return pd.DataFrame(result)


def value_ratio_distribution_by_org(data: pd.DataFrame, features: List[str], 
                                     oos_orgs: list = None, 
                                     value_bins: list = [0, 0.15, 0.35, 0.65, 0.95, 1.0]) -> pd.DataFrame:
    """统计每个机构下不同有值率区间的变量个数和占比
    
    Args:
        data: 数据（包含new_org列）
        features: 特征列表
        oos_orgs: 贷外机构列表
        value_bins: 有值率区间边界 [0, 0.15, 0.35, 0.65, 0.95, 1.0]
    
    Returns:
        有值率分布统计表
    """
    if oos_orgs is None:
        oos_orgs = []
    
    # 定义区间标签
    bin_labels = ['[0, 15%)', '[15%, 35%)', '[35%, 65%)', '[65%, 95%)', '[95%, 100%]']
    
    result = []
    
    # 获取所有机构
    orgs = data['new_org'].unique()
    
    for org in orgs:
        org_data = data[data['new_org'] == org]
        
        # 判断机构类型
        org_type = '贷外' if org in oos_orgs else '建模'
        
        # 计算每个变量的有值率（非NaN的占比）
        value_ratios = {}
        for f in features:
            if f in org_data.columns:
                non_null_count = org_data[f].notna().sum()
                total_count = len(org_data)
                value_ratios[f] = non_null_count / total_count if total_count > 0 else 0
        
        # 统计各区间的变量个数
        total_vars = len(value_ratios)
        for i in range(len(value_bins) - 1):
            lower = value_bins[i]
            upper = value_bins[i + 1]
            if upper == 1.0:
                count = sum(1 for v in value_ratios.values() if lower <= v <= upper)
            else:
                count = sum(1 for v in value_ratios.values() if lower <= v < upper)
            ratio = count / total_vars if total_vars > 0 else 0
            result.append({
                '机构': org,
                '类型': org_type,
                '有值率区间': bin_labels[i],
                '变量个数': count,
                '占比': f'{ratio:.2%}'
            })
    
    return pd.DataFrame(result)


def calculate_iv_by_org(data: pd.DataFrame, features: List[str], 
                        n_jobs: int = 4) -> Tuple[pd.DataFrame, pd.DataFrame]:
    """计算分机构和整体的IV
    
    Returns:
        iv_by_org: 分机构IV明细
        iv_overall: 整体IV
    """
    from references.func import calculate_iv
    
    orgs = data['new_org'].unique()
    
    # 整体IV
    iv_overall = calculate_iv(data, features, n_jobs=n_jobs)
    iv_overall['类型'] = '整体'
    
    # 分机构IV
    iv_by_org = []
    for org in orgs:
        org_data = data[data['new_org'] == org]
        org_iv = calculate_iv(org_data, features, n_jobs=1)  # 单机构用单进程
        if len(org_iv) > 0:  # 只添加非空结果
            org_iv['机构'] = org
            org_iv['类型'] = '分机构'
            iv_by_org.append(org_iv)
    
    iv_by_org = pd.concat(iv_by_org, ignore_index=True) if iv_by_org else pd.DataFrame(columns=['变量', 'IV', '机构', '类型'])
    
    return iv_by_org, iv_overall


def calculate_psi_detail(data: pd.DataFrame, features: List[str],
                         max_psi: float = 0.1, min_months_unstable: int = 3,
                         min_sample: int = 100, n_jobs: int = 4) -> tuple:
    """计算每个变量在每个机构下的逐月PSI明细，并标记是否剔除
    
    Returns:
        data: 剔除后的数据
        dropped: 被剔除的变量汇总
        psi_summary: 完整的PSI明细（包含是否剔除标识）
    """
    orgs = data['new_org'].unique()
    
    # 构建任务
    tasks = []
    for org in orgs:
        org_data = data[data['new_org'] == org]
        months = sorted(org_data['new_date_ym'].unique())
        
        if len(months) < 2:
            continue
        
        for i in range(len(months) - 1):
            train_month = months[i]
            test_month = months[i + 1]
            
            train_data = org_data[org_data['new_date_ym'] == train_month]
            test_data = org_data[org_data['new_date_ym'] == test_month]
            
            train_n = len(train_data)
            test_n = len(test_data)
            
            for f in features:
                tasks.append((org, train_month, test_month, train_n, test_n, f, data, min_sample))
    
    # 多进程计算
    print(f"   PSI计算: {len(tasks)}个任务, 使用{n_jobs}进程")
    results = Parallel(n_jobs=n_jobs, verbose=0)(delayed(_calc_single_psi)(task) for task in tasks)
    
    psi_df = pd.DataFrame(results)
    
    if len(psi_df) == 0:
        return data, pd.DataFrame(columns=['变量', '机构数', '不稳定机构数', '原因']), pd.DataFrame(columns=['变量', '机构数', '不稳定机构数', '是否剔除', '去除条件'])
    
    # 筛选有效计算的记录
    valid_psi = psi_df[psi_df['有效计算'] == 1].copy()
    
    if len(valid_psi) == 0:
        return data, pd.DataFrame(columns=['变量', '机构数', '不稳定机构数', '原因']), pd.DataFrame(columns=['变量', '机构数', '不稳定机构数', '是否剔除', '去除条件'])
    
    # 标记不稳定
    valid_psi['不稳定'] = (valid_psi['PSI'] > max_psi).astype(int)
    
    # 汇总：每个机构每个变量的不稳定月份数
    org_summary = valid_psi.groupby(['机构', '变量'])['不稳定'].sum().reset_index()
    org_summary.columns = ['机构', '变量', '不稳定月份数']
    
    # 机构层面汇总：超过min_months_unstable个月不稳定的变量
    org_count = len(orgs)
    channel_summary = org_summary.groupby('变量').apply(
        lambda x: pd.Series({
            '机构数': org_count,
            '不稳定机构数': (x['不稳定月份数'] >= min_months_unstable).sum()
        })
    ).reset_index()
    
    # 标记需要删除的变量（超过1/3机构不稳定）
    channel_summary['需剔除'] = channel_summary['不稳定机构数'] > (channel_summary['机构数'] / 3)
    channel_summary['是否剔除'] = channel_summary['需剔除'].astype(int)
    channel_summary['去除条件'] = channel_summary.apply(
        lambda x: f'在{org_count}个机构中有超过1/3机构连续{min_months_unstable}月PSI>{max_psi}' if x['需剔除'] else '', axis=1
    )
    
    # 筛选要剔除的变量
    if len(channel_summary) > 0 and '变量' in channel_summary.columns:
        to_drop_vars = channel_summary[channel_summary['需剔除']]['变量'].tolist()
        data = data.drop(columns=[c for c in to_drop_vars if c in data.columns])
    
    # 整理剔除信息（只返回被剔除的变量）
    dropped = channel_summary[channel_summary['需剔除']].copy()
    dropped['原因'] = f'在{org_count}个机构中有超过1/3机构连续{min_months_unstable}月PSI>{max_psi}'
    
    return data, dropped[['变量', '机构数', '不稳定机构数', '原因']], channel_summary[['变量', '机构数', '不稳定机构数', '是否剔除', '去除条件']]


def export_cleaning_report(filepath: str, steps: list, 
                           iv_detail: pd.DataFrame = None,
                           iv_process: pd.DataFrame = None,
                           psi_detail: pd.DataFrame = None,
                           psi_process: pd.DataFrame = None,
                           params: dict = None,
                           iv_distribution: pd.DataFrame = None,
                           psi_distribution: pd.DataFrame = None,
                           value_ratio_distribution: pd.DataFrame = None):
    """导出清洗报告到xlsx - 每个sheet一个步骤
    
    Args:
        filepath: 输出路径
        steps: 清洗步骤列表 [(步骤名, DataFrame), ...]
        iv_detail: IV明细（每个变量在每个机构上以及整体上的IV值）
        iv_process: IV处理表（不满足设定条件的变量）
        psi_detail: PSI明细（每个变量在每个机构每个月上的PSI值）
        psi_process: PSI处理表（不满足设定条件的变量）
        params: 超参数字典，用于动态生成依据条件
        iv_distribution: IV分布统计表
        psi_distribution: PSI分布统计表
        value_ratio_distribution: 有值率分布统计表
    """
    from openpyxl import load_workbook
    
    try:
        wb = load_workbook(filepath)
    except:
        wb = Workbook()
        wb.remove(wb.active)
    
    # 汇总sheet - 只显示真正的筛选步骤
    if '汇总' in wb.sheetnames:
        del wb['汇总']
    ws = wb.create_sheet('汇总', 0)
    ws['A1'] = '数据清洗报告'
    ws['A2'] = f'生成时间: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}'
    ws['A4'] = '步骤'
    ws['B4'] = '操作明细'
    ws['C4'] = '操作结果'
    ws['D4'] = '依据条件'
    
    # 只显示真正的筛选步骤（不包含明细和分布统计）
    filter_steps = [
        'Step4-异常月份处理', 'Step6-高缺失率处理', 'Step7-IV处理', 
        'Step8-PSI处理', 'Step9-null importance处理', 'Step10-高相关性剔除'
    ]
    
    # 需要排除的步骤（明细和分布统计）
    exclude_steps = [
        'Step7-IV明细', 'Step7-IV分布统计', 'Step8-PSI明细', 
        'Step8-PSI分布统计', 'Step5-有值率分布统计'
    ]
    
    # 需要显示删除变量数量的步骤
    show_drop_count_steps = ['分离OOS数据']
    
    # 只显示参数标准的步骤（不显示操作结果）
    show_param_only_steps = ['机构样本统计', '缺失率明细']
    
    # 添加说明：各步骤独立执行
    ws['A3'] = '说明：各筛选步骤独立执行，不删除数据，仅统计不满足条件的变量'
    
    # 获取参数，如果没有传入则使用默认值
    if params is None:
        params = {}
    
    min_ym_bad_sample = params.get('min_ym_bad_sample', 10)
    min_ym_sample = params.get('min_ym_sample', 500)
    missing_ratio = params.get('missing_ratio', 0.6)
    overall_iv_threshold = params.get('overall_iv_threshold', 0.1)
    org_iv_threshold = params.get('org_iv_threshold', 0.1)
    max_org_threshold = params.get('max_org_threshold', 2)
    psi_threshold = params.get('psi_threshold', 0.1)
    max_months_ratio = params.get('max_months_ratio', 1/3)
    max_orgs = params.get('max_orgs', 4)
    gain_threshold = params.get('gain_threshold', 50)
    
    step_num = 1
    for name, df in steps:
        # 跳过明细和分布统计步骤
        if name in exclude_steps:
            continue
        
        # 去掉操作明细中的StepX-前缀
        display_name = name.replace('Step4-', '').replace('Step6-', '').replace('Step7-', '').replace('Step8-', '').replace('Step9-', '').replace('Step10-', '')
        
        # 只显示参数标准的步骤（不显示操作结果）
        if name in show_param_only_steps:
            ws.cell(4+step_num, 1, step_num)
            ws.cell(4+step_num, 2, display_name)
            result = ''
            # 依据条件：显示参数标准
            if name == '机构样本统计':
                condition = '统计各机构的样本数量和坏样本率'
            elif name == '缺失率明细':
                condition = '计算各变量的缺失率'
            else:
                condition = ''
            ws.cell(4+step_num, 3, result)
            ws.cell(4+step_num, 4, condition)
            step_num += 1
        # 显示删除变量数量的步骤
        elif name in show_drop_count_steps:
            ws.cell(4+step_num, 1, step_num)
            ws.cell(4+step_num, 2, display_name)
            if df is not None and len(df) > 0:
                if name == '分离OOS数据':
                    # 特殊处理：显示OOS和建模样本数量
                    if '变量' in df.columns and '数量' in df.columns:
                    
                        oos_count = df[df['变量'] == 'OOS样本']['数量'].values[0] if len(df[df['变量'] == 'OOS样本']) > 0 else 0
                        model_count = df[df['变量'] == '建模样本']['数量'].values[0] if len(df[df['变量'] == '建模样本']) > 0 else 0
                        result = f'OOS样本{oos_count}条，建模样本{model_count}条'
                    else:
                        result = f'{len(df)}条'
                elif '变量' in df.columns:
                    result = f'删除{len(df)}个变量'
                else:
                    result = f'删除{len(df)}个'
                condition = ''
            else:
                result = '空'
                condition = ''
            ws.cell(4+step_num, 3, result)
            ws.cell(4+step_num, 4, condition)
            step_num += 1
        elif name in filter_steps:
            ws.cell(4+step_num, 1, step_num)
            ws.cell(4+step_num, 2, display_name)
            
            # 生成操作结果和依据条件
            if df is not None and len(df) > 0:
                if name == 'Step4-异常月份处理':
                    # 操作结果：删除的月份
                    if '年月' in df.columns:
                        result = '删除' + ','.join(df['年月'].astype(str).tolist())
                    else:
                        result = '删除' + ','.join(df.iloc[:, 0].astype(str).tolist())
                    # 依据条件：参数标准
                    condition = f'坏样本数小于{min_ym_bad_sample}或总样本数小于{min_ym_sample}的月份将被剔除（独立执行）'
                elif name == 'Step6-高缺失率处理':
                    # 操作结果：删除的变量数量
                    if '变量' in df.columns:
                        result = f'删除{len(df)}个变量'
                    else:
                        result = f'删除{len(df)}个'
                    # 依据条件：参数标准
                    condition = f'整体缺失率大于{missing_ratio}的变量将被剔除（独立执行）'
                elif name == 'Step7-IV处理':
                    # 操作结果：删除的变量数量
                    if '变量' in df.columns:
                        result = f'删除{len(df)}个变量'
                    else:
                        result = f'删除{len(df)}个'
                    # 依据条件：参数标准
                    condition = f'整体IV小于{overall_iv_threshold}或在{max_org_threshold}个及以上机构上IV小于{org_iv_threshold}的变量将被剔除（独立执行）'
                elif name == 'Step8-PSI处理':
                    # 操作结果：删除的变量数量
                    if '变量' in df.columns:
                        result = f'删除{len(df)}个变量'
                    else:
                        result = f'删除{len(df)}个'
                    # 依据条件：参数标准
                    condition = f'PSI阈值{psi_threshold}，若某机构超过{max_months_ratio:.0%}月份PSI大于{psi_threshold}则该机构不稳定，若超过{max_orgs}个机构不稳定则剔除该变量（独立执行）'
                elif name == 'Step9-null importance处理':
                    # 操作结果：删除的变量数量
                    if '变量' in df.columns:
                        result = f'删除{len(df[df["状态"]=="去除"])}个变量'
                    else:
                        result = f'删除{len(df)}个'
                    # 依据条件：参数标准
                    condition = f'反转前后gain值差别绝对值小于{gain_threshold}的变量将被判定为噪音并剔除（独立执行）'
                elif name == 'Step10-高相关性剔除':
                    # 操作结果：删除的变量数量
                    if '变量' in df.columns:
                        result = f'删除{len(df)}个变量'
                    else:
                        result = f'删除{len(df)}个'
                    # 依据条件：参数标准
                    max_corr = params.get('max_corr', 0.9)
                    top_n_keep = params.get('top_n_keep', 20)
                    condition = f'相关性大于{max_corr}的变量将被剔除，保留原始gain排名前{top_n_keep}的变量（独立执行）'
                else:
                    result = '删除' + str(len(df)) + '个'
                    condition = ''
            else:
                result = '空'
                condition = ''
            
            ws.cell(4+step_num, 3, result)
            ws.cell(4+step_num, 4, condition)
            step_num += 1
    
    # 计算最终去除的变量总数（取各步骤去除变量的并集）
    all_dropped_vars = set()
    for name, df in steps:
        if name in filter_steps and df is not None and len(df) > 0 and '变量' in df.columns:
            if name == 'Step9-null importance处理':
                # null importance处理需要筛选状态为"去除"的变量
                dropped_vars = df[df['状态'] == '去除']['变量'].tolist()
            else:
                dropped_vars = df['变量'].tolist()
            # 取并集（去重）
            all_dropped_vars = all_dropped_vars.union(set(dropped_vars))
    
    # 添加最终统计行
    final_step_num = step_num
    ws.cell(4+final_step_num, 1, final_step_num)
    ws.cell(4+final_step_num, 2, '最终去除变量统计')
    ws.cell(4+final_step_num, 3, f'共去除{len(all_dropped_vars)}个变量（取各步骤去除变量的并集）')
    ws.cell(4+final_step_num, 4, '各步骤独立执行，最终去除变量为各步骤去除变量的并集')
    
    # 各步骤详情（按步骤递进顺序创建）
    # 定义sheet创建顺序
    sheet_order = [
        '机构样本统计', '分离OOS数据', 'Step4-异常月份处理', '缺失率明细',
        'Step5-有值率分布统计', 'Step6-高缺失率处理', 'Step7-IV明细', 'Step7-IV处理',
        'Step7-IV分布统计', 'Step8-PSI明细', 'Step8-PSI处理', 'Step8-PSI分布统计',
        'Step9-null importance处理', 'Step10-高相关性剔除'
    ]
    
    # 按顺序创建sheet
    for sheet_name in sheet_order:
        # 在steps中查找对应的DataFrame
        df = None
        for name, step_df in steps:
            if name == sheet_name:
                df = step_df
                break
        
        if df is not None:
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
            ws_detail = wb.create_sheet(sheet_name)
            
            for j, col in enumerate(df.columns):
                ws_detail.cell(1, j+1, col)
            
            for i, row in df.iterrows():
                for j, val in enumerate(row):
                    # 直接写入值，避免字符转义问题
                    ws_detail.cell(i+2, j+1, val if val is not None else '')
            
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True)
            for cell in ws_detail[1]:
                cell.fill = header_fill
                cell.font = header_font
    
    wb.save(filepath)
    print(f"报告已保存: {filepath}")
